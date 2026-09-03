#!/usr/bin/env python3
"""
Tests for excel_report.py.

Run directly, no test runner required:

    python tests/test_excel_report.py

The workbook-writing tests need openpyxl and are skipped without it. Everything
that decides what a NUMBER in the sheet says - vintage selection, date-aligned
benchmarks, fail-closed label statuses - is stdlib-only and always runs, because
that is the part that can quietly be wrong.
"""
from __future__ import annotations

import json
import shutil
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "claude" / "app" / "mp_v01" / "src"))
sys.path.insert(0, str(ROOT / "claude" / "app" / "mp_v01" / "tests"))

import excel_report as ex  # noqa: E402
from adapters.yahoo_daily import normalize_bars  # noqa: E402
from harness import test, run_all  # noqa: E402

try:
    import openpyxl  # noqa: F401
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False


# --- fixtures ---------------------------------------------------------------

WEEKDAYS = [
    "2024-01-01", "2024-01-02", "2024-01-03", "2024-01-04", "2024-01-05",
    "2024-01-08", "2024-01-09", "2024-01-10", "2024-01-11", "2024-01-12",
    "2024-01-15", "2024-01-16", "2024-01-17", "2024-01-18", "2024-01-19",
]


def series(dates, start=100.0, daily=0.0):
    """Raw provider rows with an exactly known compounding rate."""
    rows, px = [], start
    for d in dates:
        rows.append({"date": d, "open": px, "high": px, "low": px,
                     "close": round(px, 10), "volume": 1_000_000, "dividend": 0.0})
        px *= (1.0 + daily)
    return rows


def write_store(tmp: Path, ticker: str, dates, daily=0.0, vintage="20240101T000000Z"):
    bars = tmp / "bars"
    bars.mkdir(parents=True, exist_ok=True)
    norm = normalize_bars(series(dates, daily=daily), ticker)
    path = bars / f"{ticker}_2024-01-01_2024-01-19__v{vintage}.json"
    with open(path, "w") as f:
        json.dump({
            "ticker": ticker, "vintage_id": vintage,
            "ingested_time": datetime.now(timezone.utc).isoformat(),
            "source": "synthetic_test_fixture", "row_count": len(norm),
            "rows": norm,
        }, f, default=str)
    return path


def write_chain(tmp: Path, ticker: str, snapshot: str, contracts, vintage="20260302T160000Z"):
    d = tmp / "chains"; d.mkdir(parents=True, exist_ok=True)
    (d / f"{ticker}__v{vintage}.json").write_text(json.dumps({
        "underlying": ticker, "snapshot_time_utc": f"{snapshot}T16:05:00+00:00",
        "vintage_id": vintage, "contract_count": len(contracts),
        "contracts": contracts}))


def contract(kind="CALL", strike=100.0, expiration="2026-04-06", bid=4.0, ask=4.2,
             volume=900, oi=4200, iv=0.30, status="OK"):
    mid = round((bid + ask) / 2, 4) if status == "OK" else None
    return {"type": kind, "expiration": expiration, "strike": strike,
            "bid": bid if status == "OK" else None,
            "ask": ask if status == "OK" else None, "mid": mid,
            "volume": volume, "open_interest": oi,
            "implied_volatility": iv, "status": status}


def labels_by_date(labels):
    return {l["decision_date"]: l for l in labels}


# --- vintage selection ------------------------------------------------------

@test
def newest_vintage_wins_and_older_ones_are_left_alone():
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS, vintage="20240101T000000Z")
        newer = write_store(tmp, "SPY", WEEKDAYS, vintage="20240609T113000Z")
        found = ex.find_bar_files(tmp)
        assert list(found) == ["SPY"], found
        assert found["SPY"] == newer, f"picked {found['SPY']}, wanted {newer}"
        # The superseded vintage is still on disk. Immutability is the point.
        assert len(list((tmp / "bars").glob("*.json"))) == 2
    finally:
        shutil.rmtree(tmp)


@test
def files_without_a_vintage_marker_are_ignored():
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS)
        (tmp / "bars" / "notes.json").write_text("{}")
        assert list(ex.find_bar_files(tmp)) == ["SPY"]
    finally:
        shutil.rmtree(tmp)


# --- label correctness ------------------------------------------------------

@test
def outperforming_the_benchmark_labels_one():
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS, daily=0.0)      # flat benchmark
        write_store(tmp, "MSFT", WEEKDAYS, daily=0.01)    # +1%/day
        data = ex.collect(tmp)
        first = labels_by_date(data["labels"]["MSFT"])["2024-01-01"]
        assert first["status"] == "OK", first
        assert first["y"] == 1, first
        # 5 compounding days of +1% against a flat benchmark.
        import math
        expected = 5 * math.log(1.01)
        assert abs(first["excess_log_return"] - expected) < 1e-9, first
    finally:
        shutil.rmtree(tmp)


@test
def underperforming_the_benchmark_labels_zero():
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS, daily=0.01)
        write_store(tmp, "MSFT", WEEKDAYS, daily=0.0)
        data = ex.collect(tmp)
        first = labels_by_date(data["labels"]["MSFT"])["2024-01-01"]
        assert first["status"] == "OK", first
        assert first["y"] == 0, first
        assert first["excess_log_return"] < 0, first
    finally:
        shutil.rmtree(tmp)


@test
def benchmark_itself_is_scored_on_absolute_sign():
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS, daily=0.01)
        data = ex.collect(tmp)
        first = labels_by_date(data["labels"]["SPY"])["2024-01-01"]
        assert first["y"] == 1, first
        assert first["benchmark"] == "n/a (is benchmark)", first
        assert first["excess_log_return"] > 0, first
    finally:
        shutil.rmtree(tmp)


@test
def benchmark_is_aligned_by_date_not_by_row_position():
    """The bug this guards against scores a stock against the wrong day of SPY.

    SPY is missing 2024-01-04 here. Index alignment would silently shift the
    benchmark by one session and still produce a confident label. Date
    alignment finds the hole and refuses.
    """
    tmp = Path(tempfile.mkdtemp())
    try:
        spy_dates = [d for d in WEEKDAYS if d != "2024-01-04"]
        write_store(tmp, "SPY", spy_dates, daily=0.0)
        write_store(tmp, "MSFT", WEEKDAYS, daily=0.01)
        data = ex.collect(tmp)
        by_date = labels_by_date(data["labels"]["MSFT"])
        # 2024-01-01's forward window covers 01-02..01-08, which spans the hole.
        assert by_date["2024-01-01"]["status"] == "RETURN_GAP_UNRESOLVED", by_date["2024-01-01"]
        assert by_date["2024-01-01"]["y"] is None
        assert by_date["2024-01-01"]["usable"] is False
        # A window clear of the hole still resolves normally.
        assert by_date["2024-01-08"]["status"] == "OK", by_date["2024-01-08"]
    finally:
        shutil.rmtree(tmp)


@test
def a_gap_in_the_instrument_fails_closed_too():
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS, daily=0.0)
        # A halted session: the bar exists but carries no usable close.
        norm = normalize_bars(series(WEEKDAYS, daily=0.01), "MSFT")
        norm[3]["daily_total_return"] = None
        norm[3]["status"] = "UNKNOWN"
        (tmp / "bars" / "MSFT_x__v20240101T000000Z.json").write_text(
            json.dumps({"ticker": "MSFT", "vintage_id": "20240101T000000Z", "rows": norm},
                       default=str))
        data = ex.collect(tmp)
        by_date = labels_by_date(data["labels"]["MSFT"])
        assert by_date["2024-01-01"]["status"] == "RETURN_GAP_UNRESOLVED", by_date["2024-01-01"]
    finally:
        shutil.rmtree(tmp)


@test
def the_unresolved_tail_is_reported_not_hidden():
    """The last rows have no 5-day outcome yet. They must say so, not vanish."""
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS, daily=0.0)
        data = ex.collect(tmp)
        labels = data["labels"]["SPY"]
        assert len(labels) == len(WEEKDAYS), f"{len(labels)} labels for {len(WEEKDAYS)} bars"
        tail = labels[-5:]
        for l in tail:
            assert l["status"] == "INSUFFICIENT_FORWARD_BARS", l
            assert l["y"] is None, l
        assert labels[-6]["status"] == "OK", labels[-6]
    finally:
        shutil.rmtree(tmp)


@test
def no_benchmark_means_no_invented_benchmark():
    """Without SPY, excess-vs-SPY is undefined. Substituting one is a lie."""
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "MSFT", WEEKDAYS, daily=0.01)
        data = ex.collect(tmp)
        assert data["benchmark_present"] is False
        assert data["labels"]["MSFT"] == []
        s = data["summaries"][0]
        assert s["labels_total"] == 0 and s["base_rate_y1"] is None, s
    finally:
        shutil.rmtree(tmp)


@test
def summary_counts_match_the_underlying_rows():
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS, daily=0.0)
        data = ex.collect(tmp)
        s = {x["ticker"]: x for x in data["summaries"]}["SPY"]
        assert s["bars"] == len(WEEKDAYS), s
        assert s["first_date"] == WEEKDAYS[0] and s["last_date"] == WEEKDAYS[-1], s
        # The first bar has no prior close, so exactly one row has no return.
        assert s["usable_returns"] == len(WEEKDAYS) - 1, s
        assert s["gap_rows"] == 1, s
        assert s["labels_usable"] == s["labels_total"] - 5, s
    finally:
        shutil.rmtree(tmp)


# --- workbook ---------------------------------------------------------------

@test
def workbook_has_the_expected_sheets_and_rows():
    if not HAVE_OPENPYXL:
        print("        (skipped: openpyxl not installed)")
        return
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS, daily=0.0)
        write_store(tmp, "MSFT", WEEKDAYS, daily=0.01)
        data = ex.collect(tmp)
        out = ex.write_workbook(data, tmp / "book.xlsx")
        assert out.exists()

        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames == ["README", "Summary", "Bars_MSFT", "Bars_SPY", "Labels"], \
            wb.sheetnames

        bars = wb["Bars_SPY"]
        assert bars.max_row == len(WEEKDAYS) + 1, bars.max_row
        assert [c.value for c in bars[1]][:6] == \
            ["date", "status", "open", "high", "low", "close"]
        assert bars.freeze_panes == "A2"
        assert bars.auto_filter.ref is not None

        labels = wb["Labels"]
        assert labels.max_row == 2 * len(WEEKDAYS) + 1, labels.max_row
    finally:
        shutil.rmtree(tmp)


@test
def retunable_columns_are_live_formulas_not_baked_values():
    """The whole point of Excel here is changing the window and re-reading it."""
    if not HAVE_OPENPYXL:
        print("        (skipped: openpyxl not installed)")
        return
    tmp = Path(tempfile.mkdtemp())
    try:
        long_dates = [f"2024-{m:02d}-{d:02d}" for m in (1, 2, 3) for d in range(1, 29)]
        write_store(tmp, "SPY", long_dates, daily=0.001)
        data = ex.collect(tmp)
        out = ex.write_workbook(data, tmp / "book.xlsx")
        ws = openpyxl.load_workbook(out)["Bars_SPY"]

        assert ws["J21"].value == '=IF(COUNT(F2:F21)<20,"",AVERAGE(F2:F21))', ws["J21"].value
        assert ws["L21"].value.startswith('=IF(COUNT(I2:I21)<20,""'), ws["L21"].value
        assert "SQRT(252)" in ws["L21"].value
        assert ws["K51"].value == '=IF(COUNT(F2:F51)<50,"",AVERAGE(F2:F51))', ws["K51"].value
        # No formula before enough history exists - it would reach into the header.
        assert ws["J20"].value is None and ws["K50"].value is None
    finally:
        shutil.rmtree(tmp)


@test
def label_sheet_carries_a_percent_formula_off_the_log_return():
    if not HAVE_OPENPYXL:
        print("        (skipped: openpyxl not installed)")
        return
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS, daily=0.01)
        data = ex.collect(tmp)
        out = ex.write_workbook(data, tmp / "book.xlsx")
        ws = openpyxl.load_workbook(out)["Labels"]
        assert ws["G2"].value == '=IF(F2="","",EXP(F2)-1)', ws["G2"].value
    finally:
        shutil.rmtree(tmp)


@test
def only_the_sheets_meant_to_have_formulas_have_them():
    """Excel reported: "Removed Records: Formula from /xl/worksheets/sheet1.xml".

    sheet1 is README, which has no formulas by design — but one documentation
    line began with "=" (explaining the excess_return_pct column), so openpyxl
    wrote the sentence as a formula. Excel could not parse prose as a formula,
    stripped it on open, and reported what looks like data loss. The data was
    never involved.

    The invariant is not "README has no formulas" but the stronger one below:
    a formula cell may exist only where this exporter deliberately writes one.
    """
    if not HAVE_OPENPYXL:
        print("        (skipped: openpyxl not installed)")
        return
    import re
    import zipfile

    # Long enough that the trailing-window formulas actually appear; they start
    # at row 21 by design, so the 15-row WEEKDAYS fixture would produce none.
    long_dates = [f"2024-{m:02d}-{d:02d}" for m in (1, 2, 3) for d in range(1, 29)]

    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", long_dates, daily=0.0)
        write_store(tmp, "MSFT", long_dates, daily=0.01)
        data = ex.collect(tmp)
        out = ex.write_workbook(data, tmp / "book.xlsx")

        wb = openpyxl.load_workbook(out)
        allowed = {"Labels"} | {n for n in wb.sheetnames if n.startswith("Bars_")}
        for name in wb.sheetnames:
            found = [c.coordinate for row in wb[name].iter_rows() for c in row
                     if c.data_type == "f"]
            if name in allowed:
                assert found, f"{name} should carry retunable formulas, has none"
            else:
                assert not found, f"{name} has unintended formula cells: {found[:5]}"

        # And at the XML level, which is what Excel actually parses.
        z = zipfile.ZipFile(out)
        readme_part = z.read("xl/worksheets/sheet1.xml").decode()
        assert "<f>" not in readme_part, "README part still contains a formula record"

        # The line that caused it must survive intact as readable text.
        text = "\n".join(str(c.value) for row in wb["README"].iter_rows()
                          for c in row if c.value)
        assert "EXP(excess_log_return)-1" in text, "the explanatory line was lost"
    finally:
        shutil.rmtree(tmp)


@test
def missing_benchmark_is_written_into_the_readme_not_just_stderr():
    if not HAVE_OPENPYXL:
        print("        (skipped: openpyxl not installed)")
        return
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "MSFT", WEEKDAYS, daily=0.01)
        data = ex.collect(tmp)
        out = ex.write_workbook(data, tmp / "book.xlsx")
        text = "\n".join(
            str(c.value) for row in openpyxl.load_workbook(out)["README"].iter_rows()
            for c in row if c.value
        )
        assert "BENCHMARK MISSING" in text, text[:400]
    finally:
        shutil.rmtree(tmp)


# --- option chains ----------------------------------------------------------

@test
def option_sheets_appear_only_when_a_chain_was_fetched():
    if not HAVE_OPENPYXL:
        print("        (skipped: openpyxl not installed)")
        return
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS, daily=0.0)
        assert "Options_Summary" not in openpyxl.load_workbook(
            ex.write_workbook(ex.collect(tmp), tmp / "a.xlsx")).sheetnames
        write_chain(tmp, "SPY", "2024-01-12", [contract(strike=100.0)])
        names = openpyxl.load_workbook(
            ex.write_workbook(ex.collect(tmp), tmp / "b.xlsx")).sheetnames
        assert "Options_Summary" in names and "Options_SPY" in names, names
    finally:
        shutil.rmtree(tmp)


@test
def greeks_are_modelled_from_the_quote_and_obey_put_call_relations():
    """delta_call - delta_put == e^-qT, and gamma/vega are shared. If these
    drift the sheet is reporting something that is not Black-Scholes."""
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS, daily=0.0)   # flat, close == 100
        write_chain(tmp, "SPY", "2024-01-12", [
            contract("CALL", 100.0, "2024-02-16", 4.0, 4.2),
            contract("PUT", 100.0, "2024-02-16", 3.6, 3.8),
        ])
        rows = ex.collect(tmp)["options"]["SPY"]
        call = next(r for r in rows if r["type"] == "CALL")
        put = next(r for r in rows if r["type"] == "PUT")
        for r in (call, put):
            assert r["model_status"] == "OK", r["model_status"]
            assert r["iv_solved"] is not None and 0 < r["iv_solved"] < 5
            assert r["gamma"] > 0 and r["vega"] > 0
            assert r["theta_per_day"] < 0
        assert 0 < call["delta"] < 1 and -1 < put["delta"] < 0
        assert abs((call["delta"] - put["delta"]) - 1.0) < 0.01, (call["delta"], put["delta"])
    finally:
        shutil.rmtree(tmp)


@test
def the_underlying_is_the_last_bar_available_before_the_snapshot():
    """A bar is not consumable at its own close, so a chain snapshotted during
    a session may only see the PRIOR session. Using the snapshot day's own
    close would be a full day of lookahead in every delta."""
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS, daily=0.01)
        write_chain(tmp, "SPY", "2024-01-10", [contract(strike=100.0)])
        row = ex.collect(tmp)["options"]["SPY"][0]
        assert row["underlying_close_date"] == "2024-01-09", row["underlying_close_date"]
        assert row["underlying_close_date"] < "2024-01-10"
    finally:
        shutil.rmtree(tmp)


@test
def unmodellable_contracts_get_a_status_and_no_greeks():
    """Blank Greeks must always be explained, never filled with a substitute."""
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS, daily=0.0)
        write_chain(tmp, "SPY", "2024-01-12", [
            contract(strike=100.0, status="UNKNOWN"),                     # no quote
            contract(strike=100.0, expiration="2024-01-12"),              # expires today
            contract(strike=100.0, expiration="2023-12-01"),              # already expired
            contract(strike=100.0, expiration="2024-02-16", bid=140.0, ask=141.0),  # above spot
        ])
        rows = ex.collect(tmp)["options"]["SPY"]
        got = {r["model_status"] for r in rows}
        assert "NO_TWO_SIDED_QUOTE" in got and "EXPIRED_OR_BAD_DATE" in got, got
        assert "IV_UNSOLVABLE_FROM_MID" in got, got
        for r in rows:
            assert r["model_status"] != "OK"
            for g in ("delta", "gamma", "theta_per_day", "vega", "rho", "iv_solved"):
                assert r[g] is None, (r["model_status"], g, r[g])
    finally:
        shutil.rmtree(tmp)


@test
def the_liquidity_screen_uses_the_projects_own_risk_limits():
    tmp = Path(tempfile.mkdtemp())
    try:
        from gates.risk import RiskLimits
        lim = RiskLimits()
        write_store(tmp, "SPY", WEEKDAYS, daily=0.0)
        write_chain(tmp, "SPY", "2024-01-12", [
            contract(strike=100.0, expiration="2024-02-16", oi=lim.min_open_interest + 1,
                     volume=lim.min_daily_volume + 1),
            contract(strike=100.0, expiration="2024-02-16", oi=lim.min_open_interest - 1,
                     volume=lim.min_daily_volume + 1),
            contract(strike=100.0, expiration="2024-01-15"),   # inside min_dte
        ])
        rows = ex.collect(tmp)["options"]["SPY"]
        assert rows[0]["liquidity_screen"] == "PASS", rows[0]
        assert rows[1]["liquidity_screen"] == "FAIL" and not rows[1]["screen_open_interest"]
        assert rows[2]["liquidity_screen"] == "FAIL" and not rows[2]["screen_dte"]
    finally:
        shutil.rmtree(tmp)


@test
def the_gate_refuses_every_contract_because_a_snapshot_has_no_edge():
    """The screen narrows the chain; it does not make a pick. gates/risk.py
    fails closed on what a chain snapshot cannot contain, and the sheet must
    show that rather than implying the survivors are picks."""
    tmp = Path(tempfile.mkdtemp())
    try:
        write_store(tmp, "SPY", WEEKDAYS, daily=0.0)
        write_chain(tmp, "SPY", "2024-01-12", [
            contract(strike=100.0, expiration="2024-02-16", oi=9999, volume=9999)])
        row = ex.collect(tmp)["options"]["SPY"][0]
        assert row["liquidity_screen"] == "PASS"
        assert row["gate_decision"] == "PASS", row["gate_decision"]
        for missing in ("independent_events", "evidence_confidence",
                        "expected_edge_after_costs"):
            assert missing in row["gate_missing"], row["gate_missing"]
    finally:
        shutil.rmtree(tmp)


if __name__ == "__main__":
    sys.exit(0 if run_all("EXCEL EXPORT") else 1)
