#!/usr/bin/env python3
"""
Turn the point-in-time data store into an Excel workbook you can actually work in.

    python excel_report.py                       # every ticker in the store
    python excel_report.py --tickers SPY,MSFT    # just these
    python excel_report.py --out book.xlsx       # explicit path

Reads the immutable vintage files written by claude/app/mp_v01/fetch_data.py and
writes one .xlsx containing:

    README        what this is, when it was built, which files it came from
    Summary       one row per ticker: coverage, gaps, label base rate
    Bars_<TICKER> one sheet per ticker - OHLCV, dividends, daily total return,
                  the two point-in-time timestamps, plus live Excel formulas for
                  SMA20 / SMA50 / 20-day annualised vol so you can retune the
                  windows in the sheet without re-running Python
    Labels        the label contract v1.0 target, built by importing the real
                  labels/contract.py - not a re-implementation that can drift

WHAT THIS DELIBERATELY DOES NOT DO
Nothing here interpolates, forward-fills, or estimates. A missing bar stays
missing, breaks the return chain, and turns the affected labels UNRESOLVED
rather than being bridged. Blank cells in the Excel output are real gaps.

The labels are FORWARD-LOOKING by construction (that is what a training target
is). They are correct to use for fitting and scoring. They are not something a
decision at that row's date could have seen. Everything to the left of the
Labels sheet is point-in-time; the Labels sheet is the answer key.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

HERE = Path(__file__).resolve().parent
MP_V01_DIR = HERE / "claude" / "app" / "mp_v01"
DEFAULT_DATA_DIR = MP_V01_DIR / "data_store"
DEFAULT_OUT_DIR = HERE / "excel_out"
DEFAULT_PICKS_DIR = HERE / "picks"

# The label contract is imported, never re-derived here. If the definition of
# the target changes in the codebase, this workbook changes with it.
sys.path.insert(0, str(MP_V01_DIR / "src"))
from labels.contract import (  # noqa: E402
    BENCHMARK,
    HORIZON_TRADING_DAYS,
    LABEL_CONTRACT_VERSION,
    build_label,
)
from backtest.costs import CostModel  # noqa: E402
from gates.risk import RiskLimits, evaluate as evaluate_gate  # noqa: E402
from options.greeks import (  # noqa: E402
    greeks as bs_greeks,
    implied_volatility,
    years_to_expiry,
)

# The risk-free rate is an ASSUMPTION, not something the bar store contains.
# Override with --risk-free-rate. It moves the Greeks only slightly at short
# tenors, but it is an input the reader is entitled to know the value of, so it
# is printed on the README sheet rather than buried here.
DEFAULT_RISK_FREE_RATE = 0.04


# ---------------------------------------------------------------------------
# Loading the store
# ---------------------------------------------------------------------------

def find_bar_files(data_dir: Path) -> dict[str, Path]:
    """Newest vintage per ticker under <data_dir>/bars.

    fetch_data.py never overwrites: every run drops a new
    ``TICKER_start_end__vYYYYmmddTHHMMSSZ.json``. Older vintages are kept on
    purpose, so a report has to pick one. It picks the newest, and the README
    sheet records exactly which file that was.
    """
    bars_dir = data_dir / "bars"
    if not bars_dir.is_dir():
        return {}

    newest: dict[str, tuple[str, Path]] = {}
    for path in sorted(bars_dir.glob("*.json")):
        head, sep, vintage = path.stem.partition("__v")
        if not sep:
            continue                      # not a vintage file; ignore quietly
        ticker = head.split("_")[0].upper()
        # Vintage ids are UTC ISO-basic, so lexical order is chronological.
        if ticker not in newest or vintage > newest[ticker][0]:
            newest[ticker] = (vintage, path)
    return {t: p for t, (_, p) in sorted(newest.items())}


def load_bar_doc(path: Path) -> dict[str, Any]:
    with open(path, encoding="utf-8") as f:
        doc = json.load(f)
    doc.setdefault("rows", [])
    doc["_source_file"] = path.name
    return doc


def dated_rows(doc: dict[str, Any]) -> list[dict[str, Any]]:
    """Rows carrying a usable date, in date order.

    Rows the adapter marked UNKNOWN keep their place in the Bars sheet (a gap
    you can see is worth more than a gap you can't), but a row with no date at
    all cannot be positioned on a calendar and is dropped here.
    """
    rows = [r for r in doc.get("rows", []) if r.get("date")]
    return sorted(rows, key=lambda r: r["date"])


# ---------------------------------------------------------------------------
# Option chains
# ---------------------------------------------------------------------------

def find_chain_files(data_dir: Path) -> dict[str, Path]:
    """Newest chain snapshot per underlying under <data_dir>/chains."""
    chains_dir = data_dir / "chains"
    if not chains_dir.is_dir():
        return {}
    newest: dict[str, tuple[str, Path]] = {}
    for path in sorted(chains_dir.glob("*.json")):
        head, sep, vintage = path.stem.partition("__v")
        if not sep:
            continue
        ticker = head.split("_")[0].upper()
        if ticker not in newest or vintage > newest[ticker][0]:
            newest[ticker] = (vintage, path)
    return {t: p for t, (_, p) in sorted(newest.items())}


def underlying_as_of(rows: list[dict[str, Any]], snapshot_date: str):
    """Last close AVAILABLE at the snapshot, with the date it came from.

    Strictly before the snapshot date, not up to and including it. A daily bar
    is not consumable at its own close (see yahoo_daily.BAR_AVAILABILITY_LAG_HOURS),
    so a chain snapshotted during a session can only legitimately see the prior
    session's close. Using today's close here would be exactly the one-day
    lookahead the rest of this project exists to prevent - and it would flatter
    every moneyness and delta in the sheet.

    The cost is that Greeks are computed against a close that may be hours
    stale. That is why the date is returned alongside and written into the
    sheet: staleness you can see is not the same problem as staleness you can't.
    """
    best = None
    for r in rows:
        d = r.get("date")
        if not d or d >= snapshot_date:
            continue
        if r.get("close") is not None:
            if best is None or d > best[0]:
                best = (d, float(r["close"]))
    return best if best else (None, None)


def trailing_dividend_yield(rows: list[dict[str, Any]], as_of_date: str) -> float:
    """Trailing 12-month cash dividends over spot, from observed bars only.

    Better than assuming zero for dividend payers, and it uses data already in
    the store rather than a new source. It is still backward-looking: a cut or
    raise inside the horizon is not anticipated.
    """
    from datetime import date as _date
    try:
        end = _date.fromisoformat(as_of_date[:10])
    except (ValueError, TypeError):
        return 0.0
    start = end.replace(year=end.year - 1) if end.month != 2 or end.day != 29 else end.replace(year=end.year - 1, day=28)

    total, spot = 0.0, None
    for r in rows:
        d = r.get("date")
        if not d:
            continue
        try:
            dd = _date.fromisoformat(d[:10])
        except ValueError:
            continue
        if start <= dd < end:
            total += float(r.get("dividend") or 0.0)
        if dd < end and r.get("close") is not None:
            spot = float(r["close"])
    if not spot or spot <= 0 or total <= 0:
        return 0.0
    return total / spot


def build_option_rows(
    ticker: str,
    chain_doc: dict[str, Any],
    bar_rows: list[dict[str, Any]],
    *,
    risk_free_rate: float,
    limits: RiskLimits,
    costs: CostModel,
) -> list[dict[str, Any]]:
    """One row per contract: observed quote, solved IV, Greeks, and screens.

    Fails closed at every step. A contract without a two-sided quote, without an
    available underlying close, already expired, or whose mid implies no
    volatility at all gets a model_status and blank Greeks - never a plausible
    substitute.
    """
    snapshot = str(chain_doc.get("snapshot_time_utc") or "")[:10]
    und_date, spot = underlying_as_of(bar_rows, snapshot) if snapshot else (None, None)
    q = trailing_dividend_yield(bar_rows, snapshot) if (snapshot and spot) else 0.0

    out: list[dict[str, Any]] = []
    for c in chain_doc.get("contracts", []):
        kind = str(c.get("type", "")).upper()
        strike = c.get("strike")
        bid, ask = c.get("bid"), c.get("ask")
        mid = c.get("mid")
        expiry = c.get("expiration")

        T = years_to_expiry(snapshot, expiry) if snapshot else None
        dte = round(T * 365) if T else None

        row = {
            "ticker": ticker, "type": kind, "expiration": expiry, "dte": dte,
            "strike": strike, "underlying_close": spot, "underlying_close_date": und_date,
            "pct_from_spot": (strike / spot - 1.0) if (spot and strike) else None,
            "bid": bid, "ask": ask, "mid": mid,
            "spread": (ask - bid) if (bid is not None and ask is not None) else None,
            "relative_spread": ((ask - bid) / mid) if (bid is not None and ask is not None and mid) else None,
            "volume": c.get("volume"), "open_interest": c.get("open_interest"),
            "iv_yahoo": c.get("implied_volatility"),
            "iv_solved": None, "delta": None, "gamma": None,
            "theta_per_day": None, "vega": None, "rho": None,
            "round_trip_cost_1x": None, "round_trip_pct_of_notional": None,
        }

        if c.get("status") != "OK" or mid is None:
            row["model_status"] = "NO_TWO_SIDED_QUOTE"
        elif spot is None:
            row["model_status"] = "NO_AVAILABLE_UNDERLYING_BAR"
        elif T is None:
            row["model_status"] = "EXPIRED_OR_BAD_DATE"
        elif kind not in ("CALL", "PUT"):
            row["model_status"] = "UNKNOWN_OPTION_TYPE"
        else:
            iv = implied_volatility(float(mid), float(spot), float(strike), T,
                                    risk_free_rate, q, kind)
            if iv is None:
                # The mid sits outside what any volatility can produce - a stale
                # or crossed quote on an illiquid line, not a tradeable price.
                row["model_status"] = "IV_UNSOLVABLE_FROM_MID"
            else:
                g = bs_greeks(float(spot), float(strike), T, risk_free_rate, iv, q, kind)
                if g is None:
                    row["model_status"] = "GREEKS_UNAVAILABLE"
                else:
                    row.update(iv_solved=iv, **g.as_dict())
                    row["theta_per_day"] = row.pop("theta")
                    row["model_status"] = "OK"

        # Execution cost on one contract, from the repo's own cost model.
        if bid is not None and ask is not None:
            try:
                rt = costs.option_round_trip_cost(float(bid), float(ask), 1)
                row["round_trip_cost_1x"] = rt
                notional = float(mid) * 100 if mid else None
                row["round_trip_pct_of_notional"] = (rt / notional) if notional else None
            except ValueError:
                pass

        # Liquidity screen, using the project's own RiskLimits constants.
        row["screen_dte"] = bool(dte is not None and limits.min_dte <= dte <= limits.max_dte)
        row["screen_spread"] = bool(row["relative_spread"] is not None
                                    and row["relative_spread"] <= limits.max_relative_spread)
        row["screen_open_interest"] = bool((row["open_interest"] or 0) >= limits.min_open_interest)
        row["screen_volume"] = bool((row["volume"] or 0) >= limits.min_daily_volume)
        row["liquidity_screen"] = "PASS" if all(
            (row["screen_dte"], row["screen_spread"],
             row["screen_open_interest"], row["screen_volume"])) else "FAIL"

        # The deterministic gate, fed everything the snapshot actually knows.
        # It will not approve anything, because a chain snapshot contains no
        # evidence and no post-cost edge estimate - and the gate fails closed on
        # what it is not told. That is the point of showing it per row.
        verdict = evaluate_gate({
            "dte": dte, "relative_spread": row["relative_spread"],
            "open_interest": row["open_interest"], "daily_volume": row["volume"],
            "defined_risk": True,
        }, limits)
        row["gate_decision"] = verdict.decision.value
        row["gate_missing"] = ", ".join(
            f[len("missing:"):] for f in verdict.failed_gates if f.startswith("missing:"))
        out.append(row)
    return out


# ---------------------------------------------------------------------------
# Labels - built with the real contract, not a copy of it
# ---------------------------------------------------------------------------

def returns_by_date(rows: Iterable[dict[str, Any]]) -> dict[str, float | None]:
    return {r["date"]: r.get("daily_total_return") for r in rows if r.get("date")}


def build_labels_for(
    ticker: str,
    rows: list[dict[str, Any]],
    benchmark_returns: dict[str, float | None],
) -> list[dict[str, Any]]:
    """One label row per bar date, including the unresolved tail.

    The last few dates in any file resolve to INSUFFICIENT_FORWARD_BARS because
    their 5-day outcome has not happened yet. Those rows are kept rather than
    hidden: "we don't know yet" is a different statement from "no signal", and
    silently truncating them is how a sheet starts lying about its coverage.
    """
    out: list[dict[str, Any]] = []
    for i, row in enumerate(rows):
        decision_date = row["date"]
        forward = rows[i + 1 : i + 1 + HORIZON_TRADING_DAYS]
        inst_returns = [r.get("daily_total_return") for r in forward]
        # Benchmark returns are aligned by DATE, never by row index: two tickers
        # can have different halted/missing sessions, and index alignment would
        # quietly score a stock against the wrong day of SPY.
        bench_returns = [benchmark_returns.get(r["date"]) for r in forward]

        label = build_label(
            ticker,
            decision_date,
            inst_returns,
            bench_returns,
        )
        out.append(
            {
                "ticker": ticker,
                "decision_date": decision_date,
                "decision_time_utc": label.decision_time_utc.isoformat(),
                "horizon_trading_days": label.horizon_trading_days,
                "y": label.y,
                "excess_log_return": label.excess_log_return,
                "status": label.status.value,
                "usable": label.is_usable(),
                "benchmark": "n/a (is benchmark)" if ticker == BENCHMARK else BENCHMARK,
                "contract_version": label.contract_version,
            }
        )
    return out


def summarize(ticker: str, doc: dict[str, Any], rows: list[dict[str, Any]],
               labels: list[dict[str, Any]]) -> dict[str, Any]:
    usable = [l for l in labels if l["usable"]]
    ones = sum(1 for l in usable if l["y"] == 1)
    closes = [r.get("close") for r in rows if r.get("close") is not None]
    return {
        "ticker": ticker,
        "source_file": doc.get("_source_file", "UNKNOWN"),
        "vintage_id": doc.get("vintage_id", "UNKNOWN"),
        "ingested_time": doc.get("ingested_time", "UNKNOWN"),
        "source": doc.get("source", "UNKNOWN"),
        "bars": len(rows),
        "first_date": rows[0]["date"] if rows else "",
        "last_date": rows[-1]["date"] if rows else "",
        "last_close": closes[-1] if closes else None,
        "usable_returns": sum(1 for r in rows if r.get("daily_total_return") is not None),
        "gap_rows": sum(1 for r in rows if r.get("daily_total_return") is None),
        "labels_total": len(labels),
        "labels_usable": len(usable),
        "labels_unresolved": len(labels) - len(usable),
        "base_rate_y1": (ones / len(usable)) if usable else None,
    }


def collect(data_dir: Path, only: list[str] | None = None, *,
            risk_free_rate: float = DEFAULT_RISK_FREE_RATE,
            picks_dir: Path | None = None) -> dict[str, Any]:
    """Everything the workbook needs, with no Excel dependency in sight."""
    files = find_bar_files(data_dir)
    if only:
        wanted = {t.strip().upper() for t in only if t.strip()}
        files = {t: p for t, p in files.items() if t in wanted}

    docs = {t: load_bar_doc(p) for t, p in files.items()}
    rows = {t: dated_rows(d) for t, d in docs.items()}

    # Excess return is defined against SPY. Without SPY in the store the target
    # is undefined for everything else, and inventing a substitute benchmark
    # would silently change what the model is being asked to predict.
    benchmark_present = BENCHMARK in rows and any(
        r.get("daily_total_return") is not None for r in rows[BENCHMARK]
    )
    bench = returns_by_date(rows[BENCHMARK]) if benchmark_present else {}

    labels: dict[str, list[dict[str, Any]]] = {}
    for ticker, rs in rows.items():
        if ticker == BENCHMARK or benchmark_present:
            labels[ticker] = build_labels_for(ticker, rs, bench)
        else:
            labels[ticker] = []

    summaries = [summarize(t, docs[t], rows[t], labels[t]) for t in sorted(rows)]

    # Option chains are optional: they exist only if fetch_data.py ran --chains.
    limits, costs = RiskLimits(), CostModel()
    chain_files = find_chain_files(data_dir)
    if only:
        wanted = {t.strip().upper() for t in only if t.strip()}
        chain_files = {t: p for t, p in chain_files.items() if t in wanted}

    options: dict[str, list[dict[str, Any]]] = {}
    chain_docs: dict[str, dict[str, Any]] = {}
    for ticker, path in chain_files.items():
        with open(path, encoding="utf-8") as f:
            doc = json.load(f)
        doc["_source_file"] = path.name
        chain_docs[ticker] = doc
        options[ticker] = build_option_rows(
            ticker, doc, rows.get(ticker, []),
            risk_free_rate=risk_free_rate, limits=limits, costs=costs,
        )

    option_summaries = []
    for ticker in sorted(options):
        rs = options[ticker]
        modelled = [r for r in rs if r["model_status"] == "OK"]
        option_summaries.append({
            "underlying": ticker,
            "source_file": chain_docs[ticker].get("_source_file", "UNKNOWN"),
            "snapshot_time_utc": chain_docs[ticker].get("snapshot_time_utc", "UNKNOWN"),
            "underlying_close": rs[0]["underlying_close"] if rs else None,
            "underlying_close_date": rs[0]["underlying_close_date"] if rs else None,
            "contracts": len(rs),
            "two_sided_quotes": sum(1 for r in rs if r["model_status"] != "NO_TWO_SIDED_QUOTE"),
            "greeks_modelled": len(modelled),
            "passed_liquidity_screen": sum(1 for r in rs if r["liquidity_screen"] == "PASS"),
            "gate_paper_trade_candidates": sum(
                1 for r in rs if r["gate_decision"] == "PAPER_TRADE_CANDIDATE"),
        })

    result = {
        "generated_utc": datetime.now(timezone.utc),
        "data_dir": str(data_dir),
        "files": {t: str(p) for t, p in files.items()},
        "chain_files": {t: str(p) for t, p in chain_files.items()},
        "docs": docs,
        "rows": rows,
        "labels": labels,
        "summaries": summaries,
        "options": options,
        "option_summaries": option_summaries,
        "risk_free_rate": risk_free_rate,
        "benchmark_present": benchmark_present,
        "pick_history": {},
    }

    # The pick history is a VIEW over the frozen files, resolved against the
    # bars just loaded. It grows by accumulating files on disk, not by anything
    # the workbook remembers, so regenerating is always safe.
    pd_ = picks_dir if picks_dir is not None else DEFAULT_PICKS_DIR
    result["picks_dir"] = str(pd_)
    result["pick_history"] = load_pick_history(Path(pd_), result, data_dir)
    return result


# ---------------------------------------------------------------------------
# Excel writing (the only part that needs openpyxl)
# ---------------------------------------------------------------------------

HEADER_FILL = "1F3B4D"
BAR_COLUMNS = [
    ("date", 12, "yyyy-mm-dd"),
    ("status", 15, None),
    ("open", 11, "#,##0.0000"),
    ("high", 11, "#,##0.0000"),
    ("low", 11, "#,##0.0000"),
    ("close", 11, "#,##0.0000"),
    ("volume", 14, "#,##0"),
    ("dividend", 10, "#,##0.0000"),
    ("daily_total_return", 18, "0.0000%"),
    ("sma_20", 12, "#,##0.0000"),
    ("sma_50", 12, "#,##0.0000"),
    ("vol_20d_annualised", 19, "0.00%"),
    ("available_time_utc", 26, None),
    ("event_time_utc", 26, None),
]

LABEL_COLUMNS = [
    ("ticker", 10, None),
    ("decision_date", 14, "yyyy-mm-dd"),
    ("decision_time_utc", 26, None),
    ("horizon_trading_days", 20, "0"),
    ("y", 6, "0"),
    ("excess_log_return", 18, "0.000000"),
    ("excess_return_pct", 18, "0.0000%"),
    ("status", 28, None),
    ("usable", 9, None),
    ("benchmark", 19, None),
    ("contract_version", 17, None),
]

OPTION_COLUMNS = [
    ("type", 7, None), ("expiration", 12, "yyyy-mm-dd"), ("dte", 6, "0"),
    ("strike", 10, "#,##0.00"), ("pct_from_spot", 14, "0.0%"),
    ("underlying_close", 16, "#,##0.00"), ("underlying_close_date", 20, None),
    ("bid", 9, "#,##0.0000"), ("ask", 9, "#,##0.0000"), ("mid", 9, "#,##0.0000"),
    ("spread", 9, "#,##0.0000"), ("relative_spread", 15, "0.0%"),
    ("volume", 10, "#,##0"), ("open_interest", 13, "#,##0"),
    ("iv_yahoo", 10, "0.0%"), ("iv_solved", 10, "0.0%"),
    ("delta", 9, "0.0000"), ("gamma", 10, "0.000000"),
    ("theta_per_day", 14, "0.0000"), ("vega", 9, "0.0000"), ("rho", 9, "0.0000"),
    ("model_status", 26, None),
    ("round_trip_cost_1x", 18, "$#,##0.00"),
    ("round_trip_pct_of_notional", 25, "0.0%"),
    ("liquidity_screen", 16, None),
    ("screen_dte", 11, None), ("screen_spread", 14, None),
    ("screen_open_interest", 20, None), ("screen_volume", 14, None),
    ("gate_decision", 22, None), ("gate_missing", 46, None),
]

OPTION_SUMMARY_COLUMNS = [
    ("underlying", 12, None), ("snapshot_time_utc", 26, None),
    ("underlying_close", 16, "#,##0.00"), ("underlying_close_date", 20, None),
    ("contracts", 11, "#,##0"), ("two_sided_quotes", 17, "#,##0"),
    ("greeks_modelled", 16, "#,##0"), ("passed_liquidity_screen", 23, "#,##0"),
    ("gate_paper_trade_candidates", 27, "#,##0"), ("source_file", 40, None),
]

SUMMARY_COLUMNS = [
    ("ticker", 10, None),
    ("bars", 8, "#,##0"),
    ("first_date", 12, None),
    ("last_date", 12, None),
    ("last_close", 12, "#,##0.0000"),
    ("usable_returns", 15, "#,##0"),
    ("gap_rows", 10, "#,##0"),
    ("labels_total", 13, "#,##0"),
    ("labels_usable", 14, "#,##0"),
    ("labels_unresolved", 18, "#,##0"),
    ("base_rate_y1", 13, "0.00%"),
    ("vintage_id", 20, None),
    ("source_file", 46, None),
    ("ingested_time", 26, None),
    ("source", 26, None),
]


def _as_date(value: Any):
    """ISO date string -> real Excel date, so sorting and charting behave."""
    if isinstance(value, str) and len(value) == 10:
        try:
            return date.fromisoformat(value)
        except ValueError:
            return value
    return value


def _sheet_name(prefix: str, ticker: str) -> str:
    clean = "".join(c for c in ticker if c not in "[]:*?/\\")
    return f"{prefix}{clean}"[:31]


def _write_header(ws, columns) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor=HEADER_FILL)
    font = Font(bold=True, color="FFFFFF")
    for idx, (name, width, _fmt) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def _finish(ws, columns, n_rows: int) -> None:
    if n_rows:
        last_col = ws.cell(row=1, column=len(columns)).column_letter
        ws.auto_filter.ref = f"A1:{last_col}{n_rows + 1}"
    for idx, (_name, _width, fmt) in enumerate(columns, start=1):
        if not fmt:
            continue
        letter = ws.cell(row=1, column=idx).column_letter
        for row in range(2, n_rows + 2):
            ws[f"{letter}{row}"].number_format = fmt


def _text_cell(ws, row: int, col: int, value: Any):
    """Write a value Excel must treat as text, never as a formula.

    openpyxl infers a formula from a leading "=", so a documentation line that
    happens to start with one is written as a formula. Excel then cannot parse
    the prose that follows, strips it on open, and reports "Removed Records:
    Formula from /xl/worksheets/sheetN.xml part" — which reads like data loss
    and is not. Prose is never a formula; say so explicitly.
    """
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"
    return cell


def _write_readme(ws, data: dict[str, Any]) -> None:
    from openpyxl.styles import Alignment, Font

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 104

    lines: list[tuple[str, str]] = [
        ("MoneyPrinter", "Point-in-time market data and label export"),
        ("Generated (UTC)", data["generated_utc"].isoformat(timespec="seconds")),
        ("Data store", data["data_dir"]),
        ("Label contract", f"v{LABEL_CONTRACT_VERSION}"),
        ("Risk-free rate used", f"{data.get('risk_free_rate', DEFAULT_RISK_FREE_RATE):.4f} "
                                f"- an ASSUMPTION, not observed data. Override with "
                                f"--risk-free-rate. Affects every Greek."),
        ("", ""),
        ("READ THIS FIRST", ""),
        ("Paper/simulation only",
         "Research output. Not financial advice, not a recommendation, and no live order "
         "path exists anywhere in this project."),
        ("Nothing is filled in",
         "Missing bars are left missing. They break the return chain instead of being "
         "bridged, and they turn the affected labels unresolved. A blank cell is a real "
         "gap in the data, not a rendering artefact."),
        ("Point-in-time columns",
         "available_time_utc is the only timestamp a backtest may filter on. It is the "
         "morning AFTER the session, not the session's own close - a daily bar for date D "
         "is not consumable at 15:45 ET on D."),
        ("The Labels sheet is the answer key",
         "Labels are forward-looking by construction. Correct for fitting and scoring, "
         "never as an input feature. Everything else in this workbook is knowable in "
         "real time; the Labels sheet is not."),
        ("", ""),
        ("SHEETS", ""),
        ("Summary", "One row per ticker: coverage, gap count, and the label base rate."),
        ("Bars_<TICKER>",
         "Raw OHLCV, dividends, daily TOTAL return (includes distributions, computed from "
         "raw closes rather than vendor-adjusted ones), and both PIT timestamps."),
        ("Labels",
         f"Binary sign of the {HORIZON_TRADING_DAYS}-trading-day forward log excess total "
         f"return vs {BENCHMARK}. Decision clock is 15:45 ET on the decision date; the "
         "label is measured from the 16:00 close that follows it. The 15-minute gap is "
         "deliberate - the model cannot consume the price it is scored against."),
        ("", ""),
        ("PICK SHEETS", "Present once picks/ contains at least one frozen run."),
        ("Pick_History",
         "Every pick ever made, accumulating across runs. Sourced from the frozen files in "
         "picks/, not from anything this workbook remembers - so regenerating never loses "
         "history, and committing picks/ is what preserves it."),
        ("Pick_Justifications", "The same picks with the full rationale for each."),
        ("Pick_Performance", "Per variant over RESOLVED picks only. Open positions are "
                             "excluded rather than counted as flat - a position with no "
                             "outcome yet is not a zero."),
        ("Pick_Abstentions", "Every variant/ticker that proposed nothing, and why."),
        ("exit_trigger",
         "Which pre-registered rule closed the position, found by walking the path day by "
         "day rather than only checking the horizon: DTE_FLOOR, PROFIT_TARGET, STOP_LOSS, "
         "or TIME_STOP if none fired first."),
        ("exit_ vs horizon_ columns",
         "Two different questions, never merged. exit_* is what following the rules would "
         "have returned - path dependent. horizon_* is whether the directional call was "
         "right, measured at the label horizon regardless of how the position closed."),
        ("exit_mark_method",
         "MARKET where a chain snapshot for that date holds the exact contract - a real "
         "observed quote. MODELLED otherwise: Black-Scholes at that close using the entry "
         "IV, which assumes volatility never moved. Fetch daily with --chains to accumulate "
         "the snapshots that turn modelled marks into market ones."),
        ("integrity",
         "VOID means that file's picks no longer hash to the digest recorded when they were "
         "frozen - it was edited after generation. Those rows are shown but excluded from "
         "Pick_Performance."),
        ("", ""),
        ("OPTIONS SHEETS", "Present only if you fetched with --chains."),
        ("Options_Summary", "Per underlying: contracts snapshotted, how many carry a two-sided "
                            "quote, how many could be modelled, and how many survive the "
                            "liquidity screen."),
        ("Options_<TICKER>",
         "One row per contract: the observed quote, an implied volatility solved from the mid, "
         "the five Greeks, execution cost from the project's own cost model, and the screens."),
        ("Observed vs modelled",
         "OBSERVED: bid, ask, strike, expiration, volume, open interest, and the underlying "
         "close. MODELLED: iv_solved and every Greek. A Greek is not a measurement - it is "
         "the output of a model whose assumptions (lognormal returns, constant volatility, "
         "European exercise) are all false to some degree for a listed US equity option."),
        ("Why iv_solved, not iv_yahoo",
         "iv_yahoo is Yahoo's own figure from an undocumented model, rate and dividend "
         "assumption. Feeding it into these formulas would stack this model on an unknown one "
         "and call the result a Greek. iv_solved is inverted from the observed mid here, so "
         "the chain is: quote -> one documented model -> Greeks. Both columns are shown on "
         "purpose: a large gap between them is a data-quality warning about that contract."),
        ("The underlying is a prior close",
         "Greeks are computed against the last bar AVAILABLE at the snapshot - the previous "
         "session's close, not live spot, because a bar is not consumable at its own close. "
         "underlying_close_date shows which. A big intraday move makes every delta in the "
         "sheet stale, and staleness you can see is a different problem from staleness you "
         "cannot."),
        ("model_status", "Why a row has no Greeks. Blank Greeks are always explained here and "
                         "never filled with a plausible substitute."),
        ("gate_decision", "gates/risk.py run per contract on what a chain snapshot actually "
                          "contains. It returns PASS - meaning DO NOTHING - on every row, "
                          "because the snapshot carries no evidence count, no confidence, and "
                          "no post-cost edge estimate, and the gate fails closed on what it is "
                          "not told. gate_missing names exactly what is absent. The liquidity "
                          "screen narrows the chain; it does not make a pick."),
        ("", ""),
        ("COLUMNS YOU CAN RETUNE", ""),
        ("sma_20 / sma_50",
         "Live Excel formulas, not pasted values. Widen or narrow the AVERAGE() range and "
         "the column recalculates. Both look backwards only, so they stay honest."),
        ("vol_20d_annualised",
         "STDEV of daily total return over the trailing 20 rows, times SQRT(252). Change "
         "252 if you want a different annualisation convention."),
        ("excess_return_pct",
         "=EXP(excess_log_return)-1, so you can read the target in percent instead of logs."),
        ("", ""),
        ("LABEL STATUS VALUES", ""),
        ("OK", "Resolved. y is 1 or 0 and is safe to train or score against."),
        ("INSUFFICIENT_FORWARD_BARS",
         "The 5-day outcome has not happened yet, or the file ends first. Expect this on "
         "the last few dates of every ticker. It means 'not known yet', not 'no signal'."),
        ("RETURN_GAP_UNRESOLVED",
         "A missing session sits inside the forward window, for this ticker or for the "
         "benchmark. Failed closed rather than bridged."),
        ("CORPORATE_ACTION_UNRESOLVED",
         "An unresolved split/merger makes the return series untrustworthy."),
        ("", ""),
        ("REGENERATING THIS FILE", ""),
        ("1. Fetch", "python claude/app/mp_v01/fetch_data.py --tickers SPY,QQQ,MSFT"),
        ("2. Export", "python excel_report.py"),
        ("or", "python gui.py, then Fetch data -> Build Excel workbook"),
        ("Note", "Every export writes a NEW timestamped file. Your edits to an existing "
                 "workbook are never overwritten."),
        ("", ""),
        ("SOURCE FILES USED", ""),
    ]
    for ticker, path in sorted(data["files"].items()):
        lines.append((ticker, path))
    if not data["files"]:
        lines.append(("(none)", "The data store was empty. Fetch data first."))
    if not data["benchmark_present"]:
        lines.append(("", ""))
        lines.append(
            (
                "BENCHMARK MISSING",
                f"{BENCHMARK} is not in the data store, so excess return vs {BENCHMARK} is "
                f"undefined and no labels were built for other tickers. Substituting a "
                f"different benchmark would change what the model is being asked to "
                f"predict, so nothing was substituted. Re-fetch including {BENCHMARK}.",
            )
        )

    bold = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)
    for r, (left, right) in enumerate(lines, start=1):
        a = _text_cell(ws, r, 1, left)
        b = _text_cell(ws, r, 2, right)
        a.alignment = wrap
        b.alignment = wrap
        if left and not right:
            a.font = Font(bold=True, size=12)
        elif left:
            a.font = bold


def _write_summary(ws, summaries: list[dict[str, Any]]) -> None:
    _write_header(ws, SUMMARY_COLUMNS)
    for r, row in enumerate(summaries, start=2):
        for c, (name, _w, _f) in enumerate(SUMMARY_COLUMNS, start=1):
            _text_cell(ws, r, c, row.get(name))
    _finish(ws, SUMMARY_COLUMNS, len(summaries))


def _write_bars(ws, rows: list[dict[str, Any]]) -> None:
    _write_header(ws, BAR_COLUMNS)
    for r, row in enumerate(rows, start=2):
        _text_cell(ws, r, 1, _as_date(row.get("date")))
        _text_cell(ws, r, 2, row.get("status", "UNKNOWN"))
        ws.cell(row=r, column=3, value=row.get("open"))
        ws.cell(row=r, column=4, value=row.get("high"))
        ws.cell(row=r, column=5, value=row.get("low"))
        ws.cell(row=r, column=6, value=row.get("close"))
        ws.cell(row=r, column=7, value=row.get("volume"))
        ws.cell(row=r, column=8, value=row.get("dividend"))
        ws.cell(row=r, column=9, value=row.get("daily_total_return"))
        # Live formulas rather than baked values: the point of putting this in
        # Excel is to be able to change the window and see the effect.
        if r >= 21:
            ws.cell(row=r, column=10,
                    value=f'=IF(COUNT(F{r-19}:F{r})<20,"",AVERAGE(F{r-19}:F{r}))')
            ws.cell(row=r, column=12,
                    value=f'=IF(COUNT(I{r-19}:I{r})<20,"",STDEV(I{r-19}:I{r})*SQRT(252))')
        if r >= 51:
            ws.cell(row=r, column=11,
                    value=f'=IF(COUNT(F{r-49}:F{r})<50,"",AVERAGE(F{r-49}:F{r}))')
        _text_cell(ws, r, 13, row.get("available_time"))
        _text_cell(ws, r, 14, row.get("event_time"))
    _finish(ws, BAR_COLUMNS, len(rows))


def _write_labels(ws, labels: list[dict[str, Any]]) -> None:
    _write_header(ws, LABEL_COLUMNS)
    for r, row in enumerate(labels, start=2):
        ws.cell(row=r, column=1, value=row["ticker"])
        ws.cell(row=r, column=2, value=_as_date(row["decision_date"]))
        ws.cell(row=r, column=3, value=row["decision_time_utc"])
        ws.cell(row=r, column=4, value=row["horizon_trading_days"])
        ws.cell(row=r, column=5, value=row["y"])
        ws.cell(row=r, column=6, value=row["excess_log_return"])
        ws.cell(row=r, column=7, value=f'=IF(F{r}="","",EXP(F{r})-1)')
        _text_cell(ws, r, 8, row["status"])
        ws.cell(row=r, column=9, value=bool(row["usable"]))
        ws.cell(row=r, column=10, value=row["benchmark"])
        ws.cell(row=r, column=11, value=row["contract_version"])
    _finish(ws, LABEL_COLUMNS, len(labels))


def _write_options(ws, rows: list[dict[str, Any]]) -> None:
    _write_header(ws, OPTION_COLUMNS)
    for r, row in enumerate(rows, start=2):
        for c, (name, _w, fmt) in enumerate(OPTION_COLUMNS, start=1):
            v = row.get(name)
            _text_cell(ws, r, c, _as_date(v) if fmt == "yyyy-mm-dd" else v)
    _finish(ws, OPTION_COLUMNS, len(rows))


def _write_options_summary(ws, summaries: list[dict[str, Any]]) -> None:
    _write_header(ws, OPTION_SUMMARY_COLUMNS)
    for r, row in enumerate(summaries, start=2):
        for c, (name, _w, _f) in enumerate(OPTION_SUMMARY_COLUMNS, start=1):
            _text_cell(ws, r, c, row.get(name))
    _finish(ws, OPTION_SUMMARY_COLUMNS, len(summaries))


PICK_HISTORY_COLUMNS = [
    ("decision_date", 14, "yyyy-mm-dd"), ("variant", 21, None), ("ticker", 8, None),
    ("direction", 10, None), ("composite_score", 15, "0.000"),
    ("contract", 24, None), ("dte_at_entry", 13, "0"),
    ("delta_at_entry", 14, "0.0000"), ("iv_at_entry", 12, "0.0%"),
    ("entry_fill", 11, "#,##0.0000"),
    ("status", 13, None), ("exit_trigger", 15, None),
    ("exit_date", 12, "yyyy-mm-dd"), ("days_held", 10, "0"),
    ("exit_price", 11, "#,##0.0000"), ("exit_mark_method", 17, None),
    ("exit_return_on_premium", 21, "0.0%"),
    ("exit_pnl_per_contract", 21, "$#,##0.00"),
    ("horizon_date", 13, "yyyy-mm-dd"), ("underlying_move_pct", 19, "0.00%"),
    ("direction_correct", 17, None), ("horizon_return_on_premium", 24, "0.0%"),
    ("horizon_mark_method", 20, None),
    ("horizon_trading_days", 20, "0"), ("profit_target_pct", 17, "0%"),
    ("stop_loss_pct", 14, "0%"), ("min_dte_exit", 13, "0"),
    ("integrity", 11, None), ("source_file", 42, None), ("detail", 52, None),
]

PICK_PERF_COLUMNS = [
    ("variant", 21, None), ("resolved", 10, "0"), ("direction_scored", 17, "0"),
    ("direction_correct", 18, "0"), ("direction_hit_rate", 19, "0.0%"),
    ("mean_return_on_premium", 23, "0.0%"), ("best_return", 12, "0.0%"),
    ("worst_return", 13, "0.0%"), ("wins", 7, "0"), ("losses", 8, "0"),
    ("modelled_marks", 15, "0"), ("exit_triggers", 44, None),
]

PICK_ABSTENTION_COLUMNS = [
    ("decision_date", 14, "yyyy-mm-dd"), ("variant", 21, None), ("ticker", 8, None),
    ("composite_score", 15, "0.000"), ("reason", 70, None),
]


def load_chains_by_date(data_dir: Path, ticker: str) -> dict[str, dict[str, Any]]:
    """EVERY chain vintage for one ticker, keyed by snapshot date.

    find_chain_files() returns only the newest, which is right for pricing today
    but useless for marking a position along a path that closed last week. Daily
    snapshots are the whole reason the fetcher writes immutable vintages, and
    this is what makes them pay off.
    """
    chains_dir = data_dir / "chains"
    if not chains_dir.is_dir():
        return {}
    out: dict[str, dict[str, Any]] = {}
    for path in sorted(chains_dir.glob(f"{ticker}__v*.json")):
        try:
            with open(path, encoding="utf-8") as f:
                doc = json.load(f)
        except (OSError, ValueError):
            continue
        day = str(doc.get("snapshot_time_utc") or "")[:10]
        if day:
            out[day] = doc           # later vintage for a day supersedes earlier
    return out


def load_pick_history(picks_dir: Path, data: dict[str, Any],
                      data_dir: Path) -> dict[str, Any]:
    """Every frozen pick file, resolved against the bars now in the store.

    The pick FILES are the durable record; this is only a view over them. So the
    history grows by accumulating files, not by anything this function
    remembers - which is what makes it safe to regenerate the workbook at will.

    A file whose picks no longer hash to their recorded digest is loaded but
    marked integrity=VOID on every row, and excluded from the performance
    summary. Reporting a tampered record silently would be worse than not
    reporting it, because the numbers would look real.
    """
    from backtest.costs import CostModel
    from strategy.picks import verify
    from strategy.resolve import resolve_pick, summarise

    if not picks_dir.is_dir():
        return {"outcomes": [], "abstentions": [], "performance": [],
                "files": [], "rationales": []}

    costs = CostModel()
    chain_cache: dict[str, dict[str, Any]] = {}
    outcomes, abstentions, rationales, files = [], [], [], []

    for path in sorted(picks_dir.glob("picks_*.json")):
        try:
            frozen = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            continue
        ok = verify(frozen)
        files.append({"file": path.name, "decision_date": frozen.get("decision_date"),
                      "generated_utc": frozen.get("generated_utc"),
                      "n_picks": frozen.get("n_picks"),
                      "integrity": "OK" if ok else "VOID"})

        for pick in frozen.get("picks", []):
            if pick.get("action") == "ABSTAIN":
                abstentions.append({
                    "decision_date": pick.get("decision_date"),
                    "variant": pick.get("variant"), "ticker": pick.get("ticker"),
                    "composite_score": pick.get("composite_score"),
                    "reason": pick.get("reason"),
                })
                continue
            ticker = pick.get("ticker")
            if ticker not in chain_cache:
                chain_cache[ticker] = load_chains_by_date(data_dir, ticker)
            o = resolve_pick(pick, data["rows"].get(ticker, []), chain_cache[ticker], costs)
            o["integrity"] = "OK" if ok else "VOID"
            o["source_file"] = path.name
            outcomes.append(o)
            rationales.append({
                "decision_date": pick.get("decision_date"),
                "variant": pick.get("variant"), "ticker": pick.get("ticker"),
                "action": pick.get("action"), "contract": o.get("contract"),
                "status": o.get("status"), "exit_trigger": o.get("exit_trigger"),
                "rationale": pick.get("rationale"),
            })

    trusted = [o for o in outcomes if o.get("integrity") == "OK"]
    return {"outcomes": outcomes, "abstentions": abstentions,
            "performance": summarise(trusted), "files": files,
            "rationales": rationales}


def _write_pick_history(ws, outcomes: list[dict[str, Any]]) -> None:
    _write_header(ws, PICK_HISTORY_COLUMNS)
    for r, o in enumerate(outcomes, start=2):
        for c, (name, _w, fmt) in enumerate(PICK_HISTORY_COLUMNS, start=1):
            v = o.get(name)
            _text_cell(ws, r, c, _as_date(v) if fmt == "yyyy-mm-dd" else v)
    _finish(ws, PICK_HISTORY_COLUMNS, len(outcomes))


def _write_pick_justifications(ws, rationales: list[dict[str, Any]]) -> None:
    """Prose gets its own sheet.

    A 1,500-character paragraph beside numeric columns makes every row tall
    enough to hide the table. Same records, different grain.
    """
    from openpyxl.styles import Alignment, Font

    cols = [("decision_date", 14), ("variant", 21), ("ticker", 8),
            ("contract", 24), ("status", 12), ("exit_trigger", 15),
            ("why this contract", 112)]
    _write_header(ws, [(n, w, None) for n, w in cols])
    for r, j in enumerate(rationales, start=2):
        vals = [_as_date(j.get("decision_date")), j.get("variant"), j.get("ticker"),
                j.get("contract"), j.get("status"), j.get("exit_trigger"),
                j.get("rationale")]
        for c, v in enumerate(vals, start=1):
            cell = _text_cell(ws, r, c, v)
            cell.alignment = Alignment(wrap_text=(c == len(vals)), vertical="top")
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2).font = Font(bold=True)
        ws.row_dimensions[r].height = max(58, 13 * (len(j.get("rationale") or "") // 105 + 2))
    if rationales:
        ws.auto_filter.ref = f"A1:G{len(rationales) + 1}"


def _write_simple(ws, columns, rows) -> None:
    _write_header(ws, columns)
    for r, row in enumerate(rows, start=2):
        for c, (name, _w, fmt) in enumerate(columns, start=1):
            v = row.get(name)
            _text_cell(ws, r, c, _as_date(v) if fmt == "yyyy-mm-dd" else v)
    _finish(ws, columns, len(rows))


DATA_SECTIONS = "data"      # bars, labels, option chains - the audit trail
PICK_SECTIONS = "picks"     # the pick record and its outcomes
ALL_SECTIONS = "all"


def write_workbook(data: dict[str, Any], out_path: Path,
                   sections: str = ALL_SECTIONS) -> Path:
    """Render collected data to .xlsx. Requires openpyxl; nothing else does.

    `sections` splits the workbook by what it is FOR, because the two halves are
    read on different schedules and at different sizes. The data half is the
    audit trail - thousands of bar rows you sort and check arithmetic against,
    rebuilt whenever you fetch. The picks half is the record - a few dozen rows
    you revisit as outcomes resolve, and which grows one run at a time.

    Writing both into one file every time meant a pick run rebuilt megabytes of
    bar sheets to append four rows, and left the picks buried behind twelve
    sheets of prices. They are separate files now, and separate folders.

    Both keep the README, because a workbook that cannot say what its own
    columns mean is a liability rather than an audit trail.
    """
    if sections not in (ALL_SECTIONS, DATA_SECTIONS, PICK_SECTIONS):
        raise ValueError(f"sections must be one of "
                         f"{ALL_SECTIONS!r}, {DATA_SECTIONS!r}, {PICK_SECTIONS!r}; "
                         f"got {sections!r}")
    want_data = sections in (ALL_SECTIONS, DATA_SECTIONS)
    want_picks = sections in (ALL_SECTIONS, PICK_SECTIONS)
    try:
        from openpyxl import Workbook
    except ImportError as e:  # pragma: no cover - depends on the local machine
        raise SystemExit(
            "This step needs openpyxl to write .xlsx files.\n"
            "    pip install openpyxl\n"
            f"(import failed: {e})"
        )

    wb = Workbook()
    _write_readme(wb.active, data)
    wb.active.title = "README"

    _write_summary(wb.create_sheet("Summary"), data["summaries"])

    if want_data:
        for ticker in sorted(data["rows"]):
            _write_bars(wb.create_sheet(_sheet_name("Bars_", ticker)), data["rows"][ticker])

        all_labels: list[dict[str, Any]] = []
        for ticker in sorted(data["labels"]):
            all_labels.extend(data["labels"][ticker])
        _write_labels(wb.create_sheet("Labels"), all_labels)

    if want_data and data.get("options"):
        _write_options_summary(wb.create_sheet("Options_Summary"),
                               data.get("option_summaries", []))
        for ticker in sorted(data["options"]):
            _write_options(wb.create_sheet(_sheet_name("Options_", ticker)),
                           data["options"][ticker])

    hist = data.get("pick_history") or {}
    if want_picks and (hist.get("outcomes") or hist.get("abstentions")):
        _write_pick_history(wb.create_sheet("Pick_History"), hist.get("outcomes", []))
        _write_pick_justifications(wb.create_sheet("Pick_Justifications"),
                                   hist.get("rationales", []))
        _write_simple(wb.create_sheet("Pick_Performance"), PICK_PERF_COLUMNS,
                      hist.get("performance", []))
        _write_simple(wb.create_sheet("Pick_Abstentions"), PICK_ABSTENTION_COLUMNS,
                      hist.get("abstentions", []))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def default_out_path(out_dir: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return out_dir / f"moneyprinter_{stamp}.xlsx"


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.strip().splitlines()[0])
    ap.add_argument("--data-dir", default=str(DEFAULT_DATA_DIR),
                    help="point-in-time store written by fetch_data.py")
    ap.add_argument("--out", default=None,
                    help="output .xlsx path (default: excel_out/moneyprinter_<timestamp>.xlsx)")
    ap.add_argument("--tickers", default=None,
                    help="comma-separated subset; default is everything in the store")
    ap.add_argument("--picks-dir", default=str(DEFAULT_PICKS_DIR),
                    help="frozen pick files to build the history sheets from")
    ap.add_argument("--sections", default=ALL_SECTIONS,
                    choices=[ALL_SECTIONS, DATA_SECTIONS, PICK_SECTIONS],
                    help="which half of the workbook to write: 'data' is the "
                         "audit trail (bars, labels, chains), 'picks' is the "
                         "pick record and its outcomes, 'all' is both")
    ap.add_argument("--risk-free-rate", type=float, default=DEFAULT_RISK_FREE_RATE,
                    help=f"annualised, for option pricing (default {DEFAULT_RISK_FREE_RATE}). "
                         f"An assumption, not observed data.")
    a = ap.parse_args(argv)

    data_dir = Path(a.data_dir).expanduser().resolve()
    only = a.tickers.split(",") if a.tickers else None

    print("=" * 70)
    print("MONEY PRINTER - Excel export")
    print("=" * 70)
    print(f"Data store : {data_dir}")

    if not (data_dir / "bars").is_dir():
        print(f"\nNo bars directory at {data_dir / 'bars'}.")
        print("Fetch some data first:")
        print("    python claude/app/mp_v01/fetch_data.py --tickers SPY,QQQ,MSFT")
        return 1

    data = collect(data_dir, only, risk_free_rate=a.risk_free_rate,
                   picks_dir=Path(a.picks_dir).expanduser())
    if not data["rows"]:
        print("\nThe data store has no bar files matching that selection.")
        print("Fetch some data first:")
        print("    python claude/app/mp_v01/fetch_data.py --tickers SPY,QQQ,MSFT")
        return 1

    for s in data["summaries"]:
        base = f"{s['base_rate_y1']:.1%}" if s["base_rate_y1"] is not None else "n/a"
        print(f"  {s['ticker']:<8} {s['bars']:>6} bars  {s['first_date']} -> {s['last_date']}"
              f"   labels {s['labels_usable']}/{s['labels_total']} usable   base rate {base}")

    for os_ in data.get("option_summaries", []):
        print(f"  {os_['underlying']:<8} {os_['contracts']:>6} contracts   "
              f"{os_['greeks_modelled']:>5} with Greeks   "
              f"{os_['passed_liquidity_screen']:>4} pass liquidity screen   "
              f"{os_['gate_paper_trade_candidates']} gate candidates")

    hist = data.get("pick_history") or {}
    if hist.get("outcomes"):
        res = sum(1 for o in hist["outcomes"] if o["status"] == "RESOLVED")
        op = sum(1 for o in hist["outcomes"] if o["status"] == "OPEN")
        print(f"\n  Pick history: {len(hist['outcomes'])} picks across "
              f"{len(hist['files'])} run(s) - {res} resolved, {op} still open")
        for row in hist.get("performance", []):
            hr = f"{row['direction_hit_rate']:.0%}" if row["direction_hit_rate"] is not None else "n/a"
            mr = f"{row['mean_return_on_premium']:+.1%}" if row["mean_return_on_premium"] is not None else "n/a"
            print(f"    {row['variant']:<22} {row['resolved']:>3} resolved   "
                  f"direction {hr:>5}   mean {mr:>7}")

    if not data["benchmark_present"]:
        print(f"\n  WARNING: {BENCHMARK} is not in the store. Excess return vs {BENCHMARK}")
        print("  is undefined without it, so no labels were built for the other tickers.")
        print(f"  Nothing was substituted. Re-fetch including {BENCHMARK}.")

    out_path = Path(a.out).expanduser() if a.out else default_out_path(DEFAULT_OUT_DIR)
    written = write_workbook(data, out_path.resolve(), sections=a.sections)

    print(f"\nWorkbook written: {written}")
    print("Every export is a new timestamped file, so edits you make are never overwritten.")
    print("=" * 70)
    return 0


if __name__ == "__main__":
    sys.exit(main())
